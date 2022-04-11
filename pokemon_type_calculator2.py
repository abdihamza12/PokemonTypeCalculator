from tkinter import *
from PIL import Image, ImageTk
import os
import openpyxl

path = 'Excel Pkdx V5.14 ugly.xlsx'

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

pkmn_list = []

directory = 'Pokemon Type Icons'  # Name of the directory where the Type icons are stored

root = Tk()  # Main Window
root.geometry('500x500')
root.resizable(width=True, height=True)
new_type_icon = []  # Empty list for the resized icons to be stored


list_frame = Frame(root)
list_frame.grid(row=0, column=5, rowspan=50, sticky='news')
real_list_frame = Frame(list_frame)
real_list_frame.grid(row=0, column=5, rowspan=60, sticky='ns')


global pkmn_type_p
pkmn_type_p = 'none'
global pkmn_type_s
pkmn_type_s = 'none'

global pkmn_window


# Resizes all the icons in the directory to a more manageable size that can be adjusted

image_size = 75
for image in os.listdir(directory):
    f = os.path.join(directory, image)
    pkmn_type = Image.open(f)
    icon_resized = pkmn_type.resize((image_size, image_size))
    new_type_icon.append(ImageTk.PhotoImage(icon_resized))


# Label(root, image=new_type_icon[2]).grid(row=0, column=0, padx=50, pady=15,  sticky='w')
# Label(root, image=new_type_icon[10]).grid(row=0, column=1, padx=50, pady=15,  sticky='e')
Label(root, text='+').grid(row=0, column=2, columnspan=1)


# GUI indicator for what Type is chosen and stored and stores the value

def prim_button_clicked(primary):
    p_icon = Label(root, image=new_type_icon[primary])
    p_icon.grid(row=1, column=0, columnspan=2, padx=50, pady=15,  sticky='we')
    global pkmn_type_p
    pkmn_type_p = list(os.listdir(directory))[primary][:-8]
    type_label_p = Label(root, text=pkmn_type_p)
    type_label_p.grid(row=1, column=0, columnspan=2, sticky='ews')

    is_filled = NORMAL

    return pkmn_type_p, is_filled


# GUI indicator for what secondary type is stored and checks if there is no secondary type

def sec_button_clicked(secondary):

    if secondary <= 17:
        s_icon = Label(root, image=new_type_icon[secondary])
        s_icon.grid(row=1, column=3, columnspan=2, padx=50, pady=15,  sticky='ew')
        global pkmn_type_s
        pkmn_type_s = list(os.listdir(directory))[secondary][:-8]
        type_label_s = Label(root, text=pkmn_type_s)
        type_label_s.grid(row=1, column=3, columnspan=2, sticky='ews')

        is_filled = NORMAL

    elif secondary == 45:
        s_icon = Label(root)
        s_icon.grid(row=1, column=3, columnspan=2, sticky='news')
        type_label_s = Label(root, text='none')
        type_label_s.grid(row=1, column=3, columnspan=2, sticky='ews')

        pkmn_type_s = 'none'

    return pkmn_type_s


def search_clicked():
    # print(str(pkmn_type_p + pkmn_type_s))

    for i in range(1, sheet_obj.max_row + 1):
        if str.lower(sheet_obj.cell(row=i, column=11).value) == pkmn_type_p:
            if str.lower(sheet_obj.cell(row=i, column=12).value) == pkmn_type_s:
                pkmn_list.append(sheet_obj.cell(row=i, column=3).value)

    if len(pkmn_list) == 0:
        print('there are no legal pokemon with that type combination')
        dne = Label(real_list_frame, text='there are no legal pokemon with that type combination')
        dne.pack()
    else:
        print(pkmn_list)

    for pokemon in pkmn_list:
        pkmn_windom = Label(real_list_frame, text=pokemon)
        pkmn_windom.pack(fill=BOTH)


    pkmn_list.clear()

    return


def clear_clicked():
    p_icon = Label(root)
    p_icon.grid(row=1, column=0, columnspan=2, sticky='news')
    s_icon = Label(root)
    s_icon.grid(row=1, column=3, columnspan=2, sticky='news')

    pkmn_list.clear()
    for widget in real_list_frame.winfo_children():
        widget.destroy()


    return print(pkmn_list, pkmn_type_p, pkmn_type_s)


p_blank_space = Label(root, padx=100, pady=75)
p_blank_space.grid(row=1, column=0, columnspan=2)    # Blank space to fill the window
s_blank_space = Label(root, padx=100, pady=75)
s_blank_space.grid(row=1, column=3, columnspan=2)


search_button = Button(root, text='Search', command=search_clicked, state=NORMAL)
clear_button = Button(root, text='Clear', command=clear_clicked)


# Primary type buttons with their respective colors and a parameter for their respective position in new_type_icon
normal_button = Button(root, text='Normal', bg='#A8A878', command=lambda: prim_button_clicked(12))
grass_button = Button(root, text='Grass', bg='#78c850', command=lambda: prim_button_clicked(9))
fire_button = Button(root, text='Fire', bg='#f08030', command=lambda: prim_button_clicked(6))
water_button = Button(root, text='Water', bg='#6890f0', command=lambda: prim_button_clicked(17))
electric_button = Button(root, text='Electric', bg='#f8d030', command=lambda: prim_button_clicked(3))
dark_button = Button(root, text='Dark', bg='#705746', command=lambda: prim_button_clicked(1))
psychic_button = Button(root, text='Psychic', bg='#f85888', command=lambda: prim_button_clicked(14))
fighting_button = Button(root, text='Fighting', bg='#c03028', command=lambda: prim_button_clicked(5))
flying_button = Button(root, text='Flying', bg='#a98ff3', command=lambda: prim_button_clicked(7))
rock_button = Button(root, text='Rock', bg='#b8a038', command=lambda: prim_button_clicked(15))
ground_button = Button(root, text='Ground', bg='#e0c068', command=lambda: prim_button_clicked(10))
ice_button = Button(root, text='Ice', bg='#98d8d8', command=lambda: prim_button_clicked(11))
bug_button = Button(root, text='Bug', bg='#a8b820', command=lambda: prim_button_clicked(0))
fairy_button = Button(root, text='Fairy', bg='#ee99ac', command=lambda: prim_button_clicked(4))
steel_button = Button(root, text='Steel', bg='#b7b7ce', command=lambda: prim_button_clicked(16))
poison_button = Button(root, text='Poison', bg='#a040a0', command=lambda: prim_button_clicked(13))
dragon_button = Button(root, text='Dragon', bg='#7038f8', command=lambda: prim_button_clicked(2))
ghost_button = Button(root, text='Ghost', bg='#705898', command=lambda: prim_button_clicked(8))

################################################################################################################

# Secondary type buttons with their respective colors and a parameter for their respective position in new_type_icon
normal_button2 = Button(root, text='Normal', bg='#A8A878', command=lambda: sec_button_clicked(12))
grass_button2 = Button(root, text='Grass', bg='#78c850', command=lambda: sec_button_clicked(9))
fire_button2 = Button(root, text='Fire', bg='#f08030', command=lambda: sec_button_clicked(6))
water_button2 = Button(root, text='Water', bg='#6890f0', command=lambda: sec_button_clicked(17))
electric_button2 = Button(root, text='Electric', bg='#f8d030', command=lambda: sec_button_clicked(3))
dark_button2 = Button(root, text='Dark', bg='#705746', command=lambda: sec_button_clicked(1))
psychic_button2 = Button(root, text='Psychic', bg='#f85888', command=lambda: sec_button_clicked(14))
fighting_button2 = Button(root, text='Fighting', bg='#c03028', command=lambda: sec_button_clicked(5))
flying_button2 = Button(root, text='Flying', bg='#a98ff3', command=lambda: sec_button_clicked(7))
rock_button2 = Button(root, text='Rock', bg='#b8a038', command=lambda: sec_button_clicked(15))
ground_button2 = Button(root, text='Ground', bg='#e0c068', command=lambda: sec_button_clicked(10))
ice_button2 = Button(root, text='Ice', bg='#98d8d8', command=lambda: sec_button_clicked(11))
bug_button2 = Button(root, text='Bug', bg='#a8b820', command=lambda: sec_button_clicked(0))
fairy_button2 = Button(root, text='Fairy', bg='#ee99ac', command=lambda: sec_button_clicked(4))
steel_button2 = Button(root, text='Steel', bg='#b7b7ce', command=lambda: sec_button_clicked(16))
poison_button2 = Button(root, text='Poison', bg='#a040a0', command=lambda: sec_button_clicked(13))
dragon_button2 = Button(root, text='Dragon', bg='#7038f8', command=lambda: sec_button_clicked(2))
ghost_button2 = Button(root, text='Ghost', bg='#705898', command=lambda: sec_button_clicked(8))
none_button = Button(root, text='None', bg='Gray', command=lambda: sec_button_clicked(45))

##################################################################################################

clear_button.grid(row=0, columnspan=5, sticky='new')
search_button.grid(row=13, columnspan=5, sticky='ews')

# Primary Button Placement
normal_button.grid(row=3, column=0, sticky='ew')
grass_button.grid(row=3, column=1, sticky='ew')

fire_button.grid(row=4, column=0, sticky='ew')
water_button.grid(row=4, column=1, sticky='ew')

electric_button.grid(row=5, column=0, sticky='ew')
dark_button.grid(row=5, column=1, sticky='ew')

psychic_button.grid(row=6, column=0, sticky='ew')
fighting_button.grid(row=6, column=1, sticky='ew')

flying_button.grid(row=7, column=0, sticky='ew')
rock_button.grid(row=7, column=1, sticky='ew')

ground_button.grid(row=8, column=0, sticky='ew')
ice_button.grid(row=8, column=1, sticky='ew')

bug_button.grid(row=9, column=0, sticky='ew')
fairy_button.grid(row=9, column=1, sticky='ew')

steel_button.grid(row=10, column=0, sticky='ew')
poison_button.grid(row=10, column=1, sticky='ew')

dragon_button.grid(row=11, column=0, sticky='ew')
ghost_button.grid(row=11, column=1, sticky='ew')

#############################################################################################

# Secondary Button Placement
normal_button2.grid(row=3, column=3, sticky='ew')
grass_button2.grid(row=3, column=4, sticky='ew')

fire_button2.grid(row=4, column=3, sticky='ew')
water_button2.grid(row=4, column=4, sticky='ew')

electric_button2.grid(row=5, column=3, sticky='ew')
dark_button2.grid(row=5, column=4, sticky='ew')

psychic_button2.grid(row=6, column=3, sticky='ew')
fighting_button2.grid(row=6, column=4, sticky='ew')

flying_button2.grid(row=7, column=3, sticky='ew')
rock_button2.grid(row=7, column=4, sticky='ew')

ground_button2.grid(row=8, column=3, sticky='ew')
ice_button2.grid(row=8, column=4, sticky='ew')

bug_button2.grid(row=9, column=3, sticky='ew')
fairy_button2.grid(row=9, column=4, sticky='ew')

steel_button2.grid(row=10, column=3, sticky='ew')
poison_button2.grid(row=10, column=4, sticky='ew')

dragon_button2.grid(row=11, column=3, sticky='ew')
ghost_button2.grid(row=11, column=4, sticky='ew')

none_button.grid(row=12, column=3, columnspan=2, sticky='ew')

root.mainloop()

